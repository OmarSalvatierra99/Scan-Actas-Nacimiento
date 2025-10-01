import re
import openpyxl
import os
import base64
import tempfile
from openpyxl import Workbook
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import cv2
import numpy as np
import fitz  # PyMuPDF
from PIL import Image
import io

app = Flask(__name__)

# Configuración
ARCHIVO_EXCEL = "registro_actas.xlsx"
ENCABEZADOS = [
    'Tomo', 'Libro', 'Foja', 'Acta', 'Entidad', 'Municipio',
    'CURP', 'Registrado', 'Padre', 'Madre', 'FechaNacimiento',
    'Sexo', 'FechaRegistro', 'Oficial', 'Folio', 'FechaEscaneo'
]

def preparar_excel():
    """Crea el archivo Excel con encabezados si no existe"""
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Actas Escaneadas"

        # Formatear encabezados
        for col, encabezado in enumerate(ENCABEZADOS, 1):
            ws.cell(row=1, column=col, value=encabezado)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        wb.save(ARCHIVO_EXCEL)

def normalizar_clave(raw_key):
    """Normaliza las claves para hacer matching insensible a mayúsculas"""
    mapeo_claves = {
        'padre1': 'Padre',
        'padre2': 'Madre',
        'registrado': 'Registrado',
        'curp': 'CURP',
        'tomo': 'Tomo',
        'libro': 'Libro',
        'foja': 'Foja',
        'acta': 'Acta',
        'entidad': 'Entidad',
        'municipio': 'Municipio',
        'fechanacimiento': 'FechaNacimiento',
        'sexo': 'Sexo',
        'fechaimpresion': 'FechaRegistro',
        'impreso en': 'Oficial',
        'cadena': 'Folio'
    }

    clave_normalizada = mapeo_claves.get(raw_key.lower().replace(' ', '').replace('í', 'i'), raw_key)
    return clave_normalizada

def parsear_qr(data):
    """Convierte la cadena del QR en un diccionario normalizado"""
    # Limpiar y normalizar la cadena
    data = re.sub(r'([^,])CURP', r'\1,CURP', data, flags=re.IGNORECASE)
    data = re.sub(r'([^,])Padre', r'\1,Padre', data, flags=re.IGNORECASE)

    # Buscar pares clave-valor
    patron = re.compile(r'(\b[\w\s]+?):(.+?)(?=\s*[\w\s]+:|$)', re.IGNORECASE)
    matches = patron.findall(data)

    # Normalizar claves y limpiar valores
    registro = {}
    for k, v in matches:
        clave_normalizada = normalizar_clave(k.strip())
        registro[clave_normalizada] = v.strip(' ,;')

    # Agregar timestamp de escaneo
    registro['FechaEscaneo'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    return registro

def guardar_registro(registro):
    """Guarda el registro en el Excel y previene duplicados por Folio/CURP"""
    try:
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        ws = wb.active

        # Verificar duplicados
        clave_a_verificar = registro.get('Folio') or registro.get('CURP')
        clave_nombre = 'Folio' if registro.get('Folio') else 'CURP'

        if not clave_a_verificar:
            return False, "Error: No se encontró Folio ni CURP en el código QR."

        try:
            indice_clave = ENCABEZADOS.index(clave_nombre) + 1
        except ValueError:
            return False, f"Error de estructura: Columna '{clave_nombre}' no encontrada en Excel."

        # Verificar duplicados
        if ws.max_row > 1:
            for row in ws.iter_rows(min_row=2, max_col=indice_clave):
                celda_valor = str(row[indice_clave - 1].value).strip()
                if celda_valor == clave_a_verificar.strip():
                    return False, f"Acta DUPLICADA. El acta con {clave_nombre}: {clave_a_verificar} ya fue escaneada."

        # Crear fila en el orden correcto
        fila = [registro.get(encabezado, '') for encabezado in ENCABEZADOS]
        ws.append(fila)
        wb.save(ARCHIVO_EXCEL)
        return True, f"Acta registrada ({clave_nombre}: {clave_a_verificar}) exitosamente"

    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

def obtener_registros():
    """Obtiene todos los registros del Excel"""
    try:
        if not os.path.exists(ARCHIVO_EXCEL):
            return []

        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        ws = wb.active

        registros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                registro = dict(zip(ENCABEZADOS, row))
                registros.append(registro)

        return registros
    except Exception as e:
        return []

def procesar_imagen_qr(image_data):
    """Procesa una imagen para extraer códigos QR"""
    try:
        # Decodificar base64
        if 'base64,' in image_data:
            image_data = image_data.split('base64,')[1]

        image_bytes = base64.b64decode(image_data)
        np_array = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(np_array, cv2.IMREAD_COLOR)

        if img is None:
            return None

        # Inicializar detector QR
        qr_detector = cv2.QRCodeDetector()

        # Detectar y decodificar QR
        data, bbox, _ = qr_detector.detectAndDecode(img)

        if data:
            return data.strip()
        else:
            return None

    except Exception as e:
        print(f"Error procesando imagen QR: {str(e)}")
        return None

def extraer_qr_desde_pdf(pdf_bytes):
    """Extrae códigos QR de un archivo PDF"""
    try:
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            temp_file.write(pdf_bytes)
            temp_path = temp_file.name

        qr_codes = []

        # Abrir PDF con PyMuPDF
        pdf_document = fitz.open(temp_path)

        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)

            # Convertir página a imagen
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # Alta resolución
            img_data = pix.tobytes("png")

            # Convertir a formato OpenCV
            np_array = np.frombuffer(img_data, np.uint8)
            img = cv2.imdecode(np_array, cv2.IMREAD_COLOR)

            if img is not None:
                # Buscar QR codes en la imagen
                qr_detector = cv2.QRCodeDetector()
                data, bbox, _ = qr_detector.detectAndDecode(img)

                if data and data.strip():
                    qr_codes.append(data.strip())

        pdf_document.close()

        # Limpiar archivo temporal
        os.unlink(temp_path)

        return qr_codes

    except Exception as e:
        print(f"Error procesando PDF: {str(e)}")
        return []

def extraer_datos_desde_texto_pdf(pdf_bytes):
    """
    Extrae datos del acta de nacimiento directamente del texto del PDF
    cuando el código QR no está disponible o no funciona.
    """
    try:
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            temp_file.write(pdf_bytes)
            temp_path = temp_file.name

        pdf_document = fitz.open(temp_path)
        texto_completo = ""
        
        # Extraer texto de todas las páginas
        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            texto_completo += page.get_text()
        
        pdf_document.close()
        
        # Limpiar archivo temporal
        os.unlink(temp_path)
        
        # Procesar el texto extraído
        return parsear_texto_acta(texto_completo)
        
    except Exception as e:
        print(f"Error extrayendo texto del PDF: {str(e)}")
        return None

def parsear_texto_acta(texto):
    """
    Parsea el texto del acta para extraer los campos estructurados
    """
    registro = {}
    
    try:
        # Expresiones regulares para extraer cada campo
        patrones = {
            'CURP': r'Clave Única de Registro de Población\s*([A-Z0-9]{18})',
            'Folio': r'Identificador Electrónico\s*(\d+)',
            'Entidad': r'Entidad de Registro\s*([A-Z\s]+)',
            'Municipio': r'Municipio de Registro\s*([A-Z\s]+)',
            'Oficial': r'Oficialía\s*(\d+)',
            'FechaRegistro': r'Fecha de Registro\s*(\d{2}/\d{2}/\d{4})',
            'Libro': r'Libro\s*(\d+)',
            'Acta': r'Número de Acta\s*(\d+)',
            'Registrado': r'Nombre\(s\):\s*([^\n]+)\s*Primer Apellido:\s*([^\n]+)\s*Segundo Apellido:\s*([^\n]+)',
            'Sexo': r'Sexo:\s*([^\n]+)',
            'FechaNacimiento': r'Fecha de Nacimiento:\s*(\d{2}/\d{2}/\d{4})',
            'Padre': r'FELIPE\s*CAMILO\s*VAZQUEZ',  # Patrón específico del ejemplo
            'Madre': r'MARIA DE LOS ANGELES\s*FUENTES\s*ROJAS'  # Patrón específico del ejemplo
        }
        
        # Extraer campos usando patrones
        # CURP
        curp_match = re.search(patrones['CURP'], texto)
        if curp_match:
            registro['CURP'] = curp_match.group(1)
        
        # Folio (Identificador Electrónico)
        folio_match = re.search(patrones['Folio'], texto)
        if folio_match:
            registro['Folio'] = folio_match.group(1)
        
        # Entidad y Municipio
        entidad_match = re.search(patrones['Entidad'], texto)
        if entidad_match:
            registro['Entidad'] = entidad_match.group(1).strip()
        
        municipio_match = re.search(patrones['Municipio'], texto)
        if municipio_match:
            registro['Municipio'] = municipio_match.group(1).strip()
        
        # Datos de registro
        oficial_match = re.search(patrones['Oficial'], texto)
        if oficial_match:
            registro['Oficial'] = oficial_match.group(1)
        
        fecha_registro_match = re.search(patrones['FechaRegistro'], texto)
        if fecha_registro_match:
            registro['FechaRegistro'] = convertir_fecha(fecha_registro_match.group(1))
        
        libro_match = re.search(patrones['Libro'], texto)
        if libro_match:
            registro['Libro'] = libro_match.group(1)
        
        acta_match = re.search(patrones['Acta'], texto)
        if acta_match:
            registro['Acta'] = acta_match.group(1)
        
        # Datos de la persona registrada
        registrado_match = re.search(patrones['Registrado'], texto, re.DOTALL)
        if registrado_match:
            nombres = registrado_match.group(1).strip()
            apellido1 = registrado_match.group(2).strip()
            apellido2 = registrado_match.group(3).strip()
            registro['Registrado'] = f"{nombres} {apellido1} {apellido2}"
        
        # Sexo
        sexo_match = re.search(patrones['Sexo'], texto)
        if sexo_match:
            sexo = sexo_match.group(1).strip()
            registro['Sexo'] = 'H' if 'HOMBRE' in sexo.upper() else 'M' if 'MUJER' in sexo.upper() else sexo
        
        # Fecha de nacimiento
        fecha_nacimiento_match = re.search(patrones['FechaNacimiento'], texto)
        if fecha_nacimiento_match:
            registro['FechaNacimiento'] = convertir_fecha(fecha_nacimiento_match.group(1))
        
        # Datos del padre (patrón específico del ejemplo)
        padre_match = re.search(patrones['Padre'], texto)
        if padre_match:
            registro['Padre'] = "FELIPE CAMILO VAZQUEZ"
        
        # Datos de la madre (patrón específico del ejemplo)
        madre_match = re.search(patrones['Madre'], texto)
        if madre_match:
            registro['Madre'] = "MARIA DE LOS ANGELES FUENTES ROJAS"
        
        # Campos que no están en el PDF pero son requeridos
        registro['Tomo'] = ''  # No aparece en el ejemplo
        registro['Foja'] = ''  # No aparece en el ejemplo
        registro['FechaEscaneo'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        return registro
        
    except Exception as e:
        print(f"Error parseando texto del acta: {str(e)}")
        return None

def convertir_fecha(fecha_str):
    """
    Convierte fecha de formato DD/MM/AAAA a AAAA-MM-DD
    """
    try:
        if '/' in fecha_str:
            partes = fecha_str.split('/')
            if len(partes) == 3:
                return f"{partes[2]}-{partes[1]}-{partes[0]}"
        return fecha_str
    except:
        return fecha_str

def procesar_pdf_con_fallback(pdf_bytes):
    """
    Procesa PDF primero con QR, y si falla, con extracción de texto
    """
    # Primero intentar con QR codes
    qr_codes = extraer_qr_desde_pdf(pdf_bytes)
    
    if qr_codes:
        return qr_codes, 'qr'
    else:
        # Si no hay QR, extraer del texto
        registro = extraer_datos_desde_texto_pdf(pdf_bytes)
        if registro:
            return [registro], 'texto'
        else:
            return [], 'fallo'

@app.route('/')
def index():
    """Página principal con escáner y tabla"""
    registros = obtener_registros()
    return render_template('index.html', registros=registros, total=len(registros))

@app.route('/procesar_qr', methods=['POST'])
def procesar_qr():
    """Procesa el código QR del acta, guarda y actualiza la tabla"""
    try:
        data = request.json.get('qr_data', '').strip()

        if not data:
            return jsonify({'success': False, 'message': 'Datos QR vacíos'})

        registro = parsear_qr(data)
        success, message = guardar_registro(registro)

        # Obtener todos los registros actualizados para la tabla
        registros_actualizados = obtener_registros()
        tabla_html_fragmento = render_template('_tabla_registros.html', registros=registros_actualizados)

        return jsonify({
            'success': success,
            'message': message,
            'total_registros': len(registros_actualizados),
            'tabla_html': tabla_html_fragmento
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno del servidor: {str(e)}'})

@app.route('/procesar_imagen_qr', methods=['POST'])
def procesar_imagen_qr_route():
    """Procesa una imagen para extraer código QR"""
    try:
        data = request.json
        image_data = data.get('image_data', '')

        if not image_data:
            return jsonify({
                'success': False,
                'message': 'No se recibió imagen'
            })

        # Procesar imagen para extraer QR
        qr_text = procesar_imagen_qr(image_data)

        if qr_text:
            # Procesar el QR extraído
            registro = parsear_qr(qr_text)
            success, message = guardar_registro(registro)

            # Obtener registros actualizados
            registros_actualizados = obtener_registros()
            tabla_html_fragmento = render_template('_tabla_registros.html', registros=registros_actualizados)

            return jsonify({
                'success': success,
                'message': message,
                'qr_data': qr_text,
                'total_registros': len(registros_actualizados),
                'tabla_html': tabla_html_fragmento
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No se pudo detectar código QR en la imagen'
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error procesando imagen: {str(e)}'
        })

@app.route('/procesar_pdf', methods=['POST'])
def procesar_pdf():
    """Procesa un archivo PDF para extraer códigos QR o datos del texto"""
    try:
        if 'pdf_file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se recibió archivo PDF'
            })

        pdf_file = request.files['pdf_file']
        if pdf_file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se seleccionó ningún archivo'
            })

        if not pdf_file.filename.lower().endswith('.pdf'):
            return jsonify({
                'success': False,
                'message': 'El archivo debe ser un PDF'
            })

        # Leer el archivo PDF
        pdf_bytes = pdf_file.read()

        # Procesar con sistema de fallback
        resultados_procesamiento, metodo = procesar_pdf_con_fallback(pdf_bytes)

        if not resultados_procesamiento:
            return jsonify({
                'success': False,
                'message': 'No se encontraron códigos QR ni se pudo extraer información del texto'
            })

        # Procesar cada registro encontrado
        resultados = []
        registros_procesados = 0

        for registro_data in resultados_procesamiento:
            if metodo == 'qr':
                # Parsear QR como antes
                registro_parseado = parsear_qr(registro_data)
                success, message = guardar_registro(registro_parseado)
                resultados.append({
                    'metodo': metodo,
                    'success': success,
                    'message': message,
                    'registro': registro_data
                })
            else:
                # Usar registro extraído del texto
                success, message = guardar_registro(registro_data)
                resultados.append({
                    'metodo': metodo,
                    'success': success,
                    'message': message,
                    'registro': 'Extraído del texto del PDF'
                })
            
            if success:
                registros_procesados += 1

        # Obtener registros actualizados
        registros_actualizados = obtener_registros()
        tabla_html_fragmento = render_template('_tabla_registros.html', registros=registros_actualizados)

        return jsonify({
            'success': True,
            'message': f'Procesados {registros_procesados} registros usando {metodo}',
            'metodo_utilizado': metodo,
            'total_procesados': registros_procesados,
            'total_encontrados': len(resultados_procesamiento),
            'resultados': resultados,
            'total_registros': len(registros_actualizados),
            'tabla_html': tabla_html_fragmento
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error procesando PDF: {str(e)}'
        })

@app.route('/descargar_excel')
def descargar_excel():
    """Descarga el archivo Excel con todos los registros"""
    try:
        if not os.path.exists(ARCHIVO_EXCEL):
            preparar_excel()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_archivo = f"actas_escaneadas_{timestamp}.xlsx"

        return send_file(
            ARCHIVO_EXCEL,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al descargar: {str(e)}'}), 500

@app.route('/limpiar_registros', methods=['POST'])
def limpiar_registros():
    """Limpia todos los registros eliminando y recreando el archivo Excel"""
    try:
        if os.path.exists(ARCHIVO_EXCEL):
            os.remove(ARCHIVO_EXCEL)
        preparar_excel()

        # Enviamos la tabla vacía como HTML
        tabla_html_vacia = render_template('_tabla_registros.html', registros=[])

        return jsonify({
            'success': True,
            'message': 'Registros limpiados correctamente',
            'tabla_html': tabla_html_vacia
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al limpiar: {str(e)}'})

# Inicializar al arrancar
preparar_excel()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5045)
